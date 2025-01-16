import tkinter as tk    #导入tkinter库
from tkinter import filedialog, messagebox #导入文件对话框和消息框
import matplotlib
matplotlib.use('TkAgg')  # 必须在导入 pyplot 之前设置后端
import matplotlib.pyplot as plt #导入matplotlib库
import numpy as np #导入numpy库
import openpyxl #导入openpyxl库
from openpyxl.styles import Alignment #导入openpyxl库中的Alignment类
import math #导入math库

class ExcelProcessorApp(tk.Tk): 
    def __init__(self):
        super().__init__()
        self.title("qpcr计算器1.0")
        self.geometry("500x600")
        self.generate_chart = tk.BooleanVar()  # 是否生成图表的选项
        self.create_widgets()

    def create_widgets(self):
        """创建界面组件"""
        # 提示标签
        self.label = tk.Label(self, text="请先阅读使用须知再操作")
        self.label.pack(pady=20)

        # 是否生成图表的复选框
        self.check_button = tk.Checkbutton(
            self, text="是否生成柱状图", variable=self.generate_chart
        )
        self.check_button.pack(pady=5)

        # 按钮：显示默认信息
        self.default_button = tk.Button(self, text="使用须知", command=self.show_default_info)
        self.default_button.pack(pady=5)

 # 按钮：生成 Excel
        self.create_excel_button = tk.Button(self, text="生成 Excel模板", command=self.generate_excel)
        self.create_excel_button.pack(pady=5)

        # 创建引物数目输入框
        self.label1 = tk.Label(self, text="请输入引物数目:")
        self.label1.pack(pady=5)

        self.entry1 = tk.Entry(self)
        self.entry1.pack(pady=5) 

        # 创建基因数目输入框
        self.label2 = tk.Label(self, text="请输入基因数目:")
        self.label2.pack(pady=5)

        self.entry2 = tk.Entry(self)
        self.entry2.pack(pady=5)
         
        # 创建DNA重复数目输入框
        self.label3 = tk.Label(self, text="请输入每个DNA重复数目:()")
        self.label3.pack(pady=5)

        # 使用 StringVar 设置默认值
        self.default_value = tk.StringVar(value="3")
        self.entry3 = tk.Entry(self, textvariable=self.default_value)
        self.entry3.pack(pady=5) 

        # 创建提交按钮
        self.submit_button = tk.Button(self, text="提交", command=self.submit_values)
        self.submit_button.pack(pady=20)

    def generate_excel(self):
            # 创建一个新的 Excel 工作簿
            workbook = openpyxl.Workbook() # 创建一个新的 Excel 工作簿
            sheet = workbook.active # 获取活动的工作表
            sheet.title = "Sheet1" # 设置工作表的标题

            # 填入第一行从第二列开始的内容
            control_group = ["对照组", "实验基因1", "实验基因2", "实验基因3", "实验基因4", "实验基因5", "实验基因6"]
            sheet.append([""] + control_group)  # 第一行，从第二个单元格开始填入

            # 填入第二行从第二列开始的内容
            primers = ["引物1", "引物2", "引物3", "引物4", "引物5", "内参"]
            for row_index, primer in enumerate(primers, start=2):  # 从第二列开始
                sheet.cell(row=row_index, column=1, value=primer)

            # 对齐单元格内容
            align = Alignment(horizontal='center', vertical='center')
            for row in sheet.iter_rows(max_row=2, min_col=2, max_col=2+len(control_group)-1):
                for cell in row:
                    cell.alignment = align

            # 保存文件
            try:
                default_path = filedialog.askdirectory(title="选择保存模板的文件夹")
                if default_path:
                    file_path = f"{default_path}/Excel模板.xlsx"
                    workbook.save(file_path)
                    messagebox.showinfo("生成成功", f"Excel 模板已成功生成并保存到：\n{file_path}")
                else:
                    messagebox.showwarning("取消保存", "未选择保存路径，文件未保存")
            except Exception as e:
                messagebox.showerror("生成失败", f"发生错误：{str(e)}")

    def show_default_info(self):
        # 显示默认信息
        default_message = (
            "使用须知:\n"
            "计算方法:2-∆∆Ct方法\n"
            "生成excel模板根据需求自行修改!请注意以下几点：\n"
            "对照组的基因在excel中必须在最左边列!\n"
            "实验组基因在excel中必须在对照组的右边列!\n"
            "作为内参的引物必须在所有引物的最下面！\n"
            "数据不要占用excel中的第一行以及A列\n"
            
            "售后联系QQ:762509971\n"
        )
        messagebox.showinfo("默认信息", default_message)

    def submit_values(self):
        # 获取用户输入的值
        time1 = self.entry1.get() #引物数目
        time2 = self.entry2.get() #基因数目
        time3 = self.entry3.get() #每个DNA重复数目

        # 校验输入是否为数字
        if not (time1.isdigit() and time2.isdigit()):
            messagebox.showerror("错误", "请输入有效的数字！")
            return
    
        # 显示用户输入的数值
        messagebox.showinfo("输入成功", f"引物数目: {time1}\n基因数目: {time2}")

        # 如果需要在其他地方使用这些值，可以存储到实例变量中
        self.time1 = int(time1) # 将输入的引物数目转换为整数
        self.time2 = int(time2) # 将输入的基因数目转换为整数
        self.time3 = int(time3) # 将输入的每个DNA重复数目转换为整数
    
        # 创建一个新窗口来显示选择的文件路径
        # 选择文件按钮
        self.btn_select_file = tk.Button(self, text="选择文件", command=self.select_file)
        self.btn_select_file.pack(pady=10)

        # 处理Excel按钮
        process_button = tk.Button(self, text="处理Excel", command=self.process_excel)
        process_button.pack()

        # 显示选择的文件路径
        self.file_path_label = tk.Label(self, text="未选择文件", wraplength=350, anchor="w", justify="left")
        self.file_path_label.pack(pady=5)

    def select_file(self): 
        """打开文件选择对话框"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        ) 
        if file_path: # 如果用户选择了文件
            self.file_path_label.config(text=f"已选择文件：{file_path}") # 更新标签文本
            self.filename = file_path
            messagebox.showinfo("文件选择成功", f"文件路径：{file_path}")  # 添加提示，确认文件选择成功
        else:
            self.file_path_label.config(text="未选择文件")  # 提示用户未选择文件
    from openpyxl import load_workbook

    def process_excel(self): # 处理Excel文件的函数
        """调用计算逻辑处理 Excel 文件"""
        if self.filename: # 如果文件名存在
            workbook = openpyxl.load_workbook(self.filename)  # 打开目标excel文件
            messagebox.showinfo('提示', f"成功打开文件：{self.filename}")
        
            try: 
                sheet = workbook.create_sheet(title='Result')
                new_sheet = workbook['Result'] #选中新表格
                sheet = workbook['Sheet1'] #选中数据所在sheet
                cDNA_total = self.time3 * self.time2 + 1 #一行引物的全部单元格数
                #第一次求差值
                for No_primers in range(2, self.time1 + 1): #引物循环，从第二行开始（不包括最底下的内参）
                    for i in range(2, cDNA_total + 1):  #遍历一行，总的基因数加1，避开第一列
                        cell1 = sheet.cell(row=No_primers, column=i)
                        cell2 = sheet.cell(row=self.time1 + 1, column=i)
                        difference = cell1.value - cell2.value
                        new_sheet = workbook['Result'] #选中新表格 
                        new_sheet.cell(row=No_primers, column=i, value=difference) #将差值写入新表格
                #第一次差值的结果中 对照组的CT值差值求均值
                for i in range(2, self.time1 + 1): 
                    new_sheet = workbook['Result'] #选中数据所在表格
                    count = 0 #计数器
                    total = 0 #求和
                    for row in new_sheet.iter_rows(min_row=i, max_row=i, min_col=2, max_col=self.time3 + 1):
                        for cell in row:
                            if cell.value is not None:
                                total += cell.value
                                count += 1
                        if count > 0:
                            average = total / count
                        else:
                            average = None
                        for a in range(0, 1): 
                            b = i + self.time1 + 1 #给均值定y轴
                            new_sheet.cell(row=b, column=1, value=average) #将均值写入新表格
                #求ΔΔCT
                new_sheet = workbook['Result'] #选中数据所在表格
                for No_primers in range(2, self.time1 + 1): #引物数循环
                    for i in range(2, cDNA_total + 1): #基因数循环
                        cell1 = new_sheet.cell(row=No_primers, column=i)
                        cir3 = No_primers + self.time1 + 1 #给均值定y轴
                        cell2 = new_sheet.cell(row=cir3, column=1)
                        difference = cell1.value - cell2.value
                        new_sheet = workbook['Result'] #选中新表格
                        row_2 = No_primers + (self.time1 + 1) * 2 #给差值定y轴
                        new_sheet.cell(row_2, column=i, value=difference)  

                #求2^(-ΔΔCT)
                for i in range(2, self.time1 + 1): #引物循环次数
                    No_primers = i + (self.time1 + 1) * 2
                    for row in new_sheet.iter_rows(min_row=No_primers, max_row=No_primers, min_col=2, max_col=cDNA_total):
                        column_index = 2 #增加列计数器
                        for cell in row:
                            if cell.value is not None:
                                result_1 = pow(2, -cell.value) #求2的负次方
                            else:
                                result_1 = None
                            b = No_primers + self.time1 + 1 #给2的负次方定y轴
                            new_sheet.cell(row=b, column=column_index, value=result_1)
                            column_index += 1 #每次单元格遍历循环，列计数器+1
                #再进行一次求均数
                for i in range(2, self.time1 + 1):  #引物循环
                    new_sheet = workbook['Result'] #选中数据所在表格
                    No_primers = i +(self.time1 + 1) * 3
                    count = 0
                    total = 0
                    for row in new_sheet.iter_rows(min_row=No_primers, max_row=No_primers, min_col=2, max_col=self.time3 + 1):
                        for cell in row:
                            if cell.value is not None:
                                total += cell.value
                                count += 1
                        if count > 0:
                            average = total / count
                        else:
                            average = None
                        for a in range(0, 1):
                            b = No_primers + self.time1 + 1 #给均值定y轴
                            new_sheet.cell(row=b, column=1, value=average)  
                #各单元格数值比内参
                for c in range(2, self.time1 + 1): #引物循环
                    No_primers = c + (self.time1 + 1) * 3
                    No_primers_2 = c + (self.time1 + 1) * 4
                    for i in range(2, cDNA_total + 1): 
                        cell1 = new_sheet.cell(row=No_primers, column=i)
                        cell2 = new_sheet.cell(row=No_primers_2, column=1)
                        ratio = cell1.value / cell2.value
                        new_sheet = workbook['Result'] #选中新表格
                        e = No_primers_2 + self.time1 + 1
                        new_sheet.cell(row=e, column=i, value=ratio)    
                #输入最后2^(-ΔΔCT)标题
                No_primers = 1 + (self.time1 + 1) * 3
                new_sheet.cell(row= No_primers, column=1, value='2^(-ΔΔCT)') 
                #输入引物标题
                for i in range(2, self.time1 + 1):
                    # 获取Sheet1中A列第i行的值
                    value = sheet.cell(row=i, column=1).value
                    # 将值写入Sheet2的B列第i行
                    new_sheet.cell(row=i, column=1, value=value)
                #输入基因标题
                for i in range(2, cDNA_total + 1):
                    # 获取Sheet1中第1行的值
                    value = sheet.cell(row=1, column=i).value
                    # 将值写入
                    new_sheet.cell(row=1, column=i, value=value)

                #合并新表格的基因标题单元格
                align = Alignment(horizontal='center', vertical='center')
                for i in range(2, cDNA_total + 1, self.time3):
                    d = i + self.time3 - 1
                    new_sheet.merge_cells(start_row=1, start_column=i, end_row=1, end_column=d)
                for i in range(1, cDNA_total):
                    new_sheet.cell(1, i).alignment = align

                #在归一化后的结果写入引物
                for i in range(2, self.time1 + 1):
                    # 获取Sheet1中A列第i行的值
                    value = sheet.cell(row=i, column=1).value
                    # 将值写入Sheet2的B列第i行
                    i = i + (self.time1 + 1)* 5
                    new_sheet.cell(row=i, column=1, value=value)

                #输入最后结果标题
                No_primers = (self.time1 + 1) * 5 + 1
                new_sheet.cell(row= No_primers, column=1, value='归一化结果')

                #输入基因标题
                for i in range(2, cDNA_total + 1):
                    # 获取Sheet1中第1行的值
                    value = sheet.cell(row=1, column=i).value
                    # 将值写入
                    d = (self.time1 + 1) * 5 + 1
                    new_sheet.cell(row=d, column=i, value=value)

                #合并新表格的基因标题单元格
                align = Alignment(horizontal='center', vertical='center')
                for i in range(2, cDNA_total, self.time3):
                    c = i + self.time3 - 1
                    d = (self.time1 + 1) * 5 + 1
                    new_sheet.merge_cells(start_row=d, start_column=i, end_row=d, end_column=c)
                for i in range(1, cDNA_total):
                    d = (self.time1 + 1) * 5 + 1
                    new_sheet.cell(d, i).alignment = align
                workbook.save(self.filename)  # 保存文件
                # 计算完成后弹出提示框
                messagebox.showinfo("计算完成", f"文件 {self.filename} 的计算已成功完成！")

            except Exception as e:
                messagebox.showerror("处理失败", f"出现错误：{str(e)}")
        else:messagebox.showerror("文件打开失败")

        if self.generate_chart.get(): # 如果勾选了生成图表
                self.draw_primer_gene_chart()  # 调用绘图函数
        else:messagebox.showerror("画图失败")
        # 更新绘图部分的代码

# 更新绘图部分的代码
    def draw_primer_gene_chart(self):
        """绘制引物与基因数据的柱状图"""
        try:
            workbook = openpyxl.load_workbook(self.filename)  # 打开目标 Excel 文件
            sheet = workbook["Result"]  # 确保文件有 Result sheet

            # 获取目标数据范围 
            start_row = 3 * self.time1 + 5
            end_row = 4 * self.time1 + 3
            start_col = 2
            end_col = self.time3 * self.time2 + 1  # cDNA_total

            # 提取基因名称
            genes = [
                sheet.cell(row=1, column=col).value
                for col in range(start_col, end_col, self.time3)
            ]

            # 提取引物名称
            primers = [
                sheet.cell(row=row, column=1).value
                for row in range(2, self.time1 + 1)
            ]

            # 准备数据
            data = []
            for row in range(start_row, end_row + 1):
                row_data = [
                    sheet.cell(row=row, column=col).value
                    for col in range(start_col, end_col)
                ]
                data.append(row_data)

           # 根据总数动态计算合理的列数
            def calculate_layout(n, aspect_ratio=1.5):
                cols = math.ceil(math.sqrt(n * aspect_ratio))  # 根据总数和长宽比计算列数
                rows = math.ceil(n / cols)  # 根据列数计算行数
                return rows, cols
            
            # 子图总数
            n_subplots = len(primers)

            # 动态计算行数和列数
            rows, cols = calculate_layout(n_subplots)

          # 设置画布大小
            fig, axes = plt.subplots(
                rows, cols, 
                figsize=(cols * 3, rows * 2.5),  # 每个子图的宽度为3，高度为2.5
                dpi=200, 
                squeeze=False
            )
            axes = axes.flatten()

            for i, primer in enumerate(primers):
                primer_data = data[i]
                mean_values = []
                std_values = []

                # 计算均值和标准差
                for j in range(len(genes)):
                    group_data = primer_data[j * self.time3 : (j + 1) * self.time3]
                    mean_values.append(np.mean(group_data))
                    std_values.append(np.std(group_data))

                # 绘制柱状图
                x = np.arange(len(genes))
                axes[i].bar(
                    x,
                    mean_values,
                    yerr=std_values,
                    capsize=5,
                    color="skyblue",
                    edgecolor="black",
                    label=primer,
                )
                axes[i].set_title(primer, fontsize=14)
                axes[i].set_xticks(x)
                axes[i].set_xticklabels(genes, fontsize=12)
                axes[i].set_ylabel("Relative Expression", fontsize=12)
                axes[i].set_xlabel("Genes", fontsize=12)
            # 隐藏多余的子图
            for i in range(len(primers), len(axes)):
                fig.delaxes(axes[i])

            fig.suptitle("基因表达柱状图", fontsize=16, y=1.02)
            plt.subplots_adjust(hspace=0.5)
            plt.tight_layout()

            # 保存图表
            save_path = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG 文件", "*.png"), ("所有文件", "*.*")],
            )
            if save_path:
                plt.savefig(save_path, format="png")
                messagebox.showinfo("成功", f"图片已保存至：{save_path}")
            else:
                messagebox.showwarning("未保存", "图片未保存，绘图已完成但未存储。")

        except Exception as e:
            messagebox.showerror("错误", f"绘图时发生错误: {str(e)}")
if __name__ == "__main__":
    app = ExcelProcessorApp()
    app.mainloop()